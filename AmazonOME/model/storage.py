"""
数据存储类 - 负责Excel文件的读写操作
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook, load_workbook


class Storage:
    """
    Excel存储管理器
    职责：读取输入链接，保存输出结果
    """

    def __init__(self, config: 'Config'):
        self.config = config

    def read_input(self) -> list[tuple[str, str, str]]:
        """
        读取输入文件（单线程调用，线程安全）
        :return:
            lisr:[(WBS No. , OEM NO. Application Vehicle &Engine Model),...]
        """

        if not self.config.input_file.exists():
            raise FileNotFoundError(f'输入文件不存在: {self.config.input_file}')

        wb = load_workbook(self.config.input_file)
        ws = wb.active

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[1]:
                WBS_No      = row[0] if len(row) > 2 else ''
                OEM_No = row[1]
                url = f'https://www.amazon.com/s?k={OEM_No}&i=automotive'
                records.append((WBS_No,url,OEM_No))

        wb.close()
        return records

    def create_output(self) -> Workbook:
        """创建输出工作簿并写入表头"""
        wb = Workbook()
        ws = wb.active
        ws.append(self.config.OUTPUT_HEADERS_)
        return wb

    def save_temp(self, wb: Workbook) -> None:
        """保存临时文件"""
        wb.save(self.config.TEMP_FILE)
        print(f'已保存临时文件')

    def save_final(self, wb: Workbook) -> None:
        """保存最终结果"""
        self.config.ensure_dirs()
        wb.save(self.config.output_file)
        print(f'结果已保存: {self.config.output_file}')

    def append_record(self, ws, data: dict) -> None:
        """
        向工作表添加一条记录

        Args:
            ws: 工作表对象
            data: 字段字典，key需与OUTPUT_HEADERS对应
        """
        row = [
            data.get('URL', ''),
            data.get('OME_No', '')
        ]
        ws.append(row)

class Storage_1:
    """
    Excel存储管理器
    职责：读取输入链接，保存输出结果
    """

    def __init__(self, config: 'Config'):
        self.config = config

    def read_input(self) -> list[tuple[str, str]]:
        """
        读取输入文件（单线程调用，线程安全）
        :return:
            lisr:[(WBS No. , OEM NO. Application Vehicle &Engine Model),...]
        """

        if not self.config.input_file_de.exists():
            raise FileNotFoundError(f'输入文件不存在: {self.config.input_file_de}')

        wb = load_workbook(self.config.input_file_de)
        ws = wb.active

        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= 2 and row[1]:
                url = row[0]
                OEM_No = row[1]
                records.append((url,OEM_No))
        wb.close()
        return records

    def create_output(self) -> Workbook:
        """创建输出工作簿并写入表头"""
        wb = Workbook()
        ws = wb.active
        ws.append(self.config.OUTPUT_HEADERS_DE)
        return wb

    def save_temp(self, wb: Workbook) -> None:
        """保存临时文件"""
        wb.save(self.config.TEMP_FILE_DE)
        print(f'已保存临时文件')

    def save_final(self, wb: Workbook) -> None:
        """保存最终结果"""
        self.config.ensure_dirs()
        wb.save(self.config.output_file_de)
        print(f'结果已保存: {self.config.output_file_de}')

    def append_record(self, ws, data: dict) -> None:
        """
        向工作表添加一条记录

        Args:
            ws: 工作表对象
            data: 字段字典，key需与OUTPUT_HEADERS对应
        """
        row = [
            data.get('URL', ''),
            data.get('OME_No', ''),
            data.get('price',''),
            data.get("About_table",''),
            data.get('item specifics',''),
            data.get('description_from_the_seller',''),
            data.get('seller_feedback',''),
            data.get('category', ''),
            data.get('picture',''),
            data.get('OEM','')
        ]
        ws.append(row)