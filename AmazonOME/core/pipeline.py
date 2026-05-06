"""
多线程爬虫入口
"""
from pathlib import Path
from config.config import Config
from core.fetcher import Fetcher,Fetcher_1
from model.storage import Storage,Storage_1
from core.multi_thread_crawler1 import MultiThreadCrawler,MultiThreadCrawler_1
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from utils.logger import get_logger
log = get_logger("Pineline")

class Pipeline:
    def run(self):
        """主函数"""
        log.info('=' * 60)
        log.info('eBay 多线程商品链接爬虫')
        log.info('模式: 生产者保存HTML -> 队列传递文件路径 -> 消费者读取解析 -> 即用即删')
        log.info('=' * 60)

        # 初始化
        config = Config()
        config.ensure_dirs()

        storage = Storage(config)

        try:
            url_list = storage.read_input()
            print(f'读取到 {len(url_list)} 个URL')
        except FileNotFoundError as e:
            print(f'错误: {e}')
            return

        fetcher = Fetcher(config)
        crawler = MultiThreadCrawler(config, fetcher, storage)

        try:
            results = crawler.run(url_list)
            print(f'\n 爬取完成！共获取 {len(results)} 条数据')
        finally:
            fetcher.close()
            print('浏览器已关闭')

        try:
            wb = load_workbook(config.output_file)
            ws = wb.active
            new_rows = []  # 收集要追加的行
            for r in results:
                ome = r.get('OME_No', '')
                # 按换行拆 URL
                for u in r.get('URL', '').splitlines():
                    if u.strip():  # 跳过空串
                        new_rows.append([u, ome])

            # 把表头以下全部清空再写
            ws.delete_rows(2, ws.max_row)
            for row in new_rows:
                ws.append(row)

            # 顺手给 URL 列自动换行（可选）
            for cell in ws['A'][1:]:
                cell.alignment = Alignment(wrap_text=True)

            wb.save(config.output_file)
            print(f'已拆成 {len(new_rows)} 行，每行一个 URL')
        except Exception as e:
            print('拆行失败:', e)

class Pipeline_1:
    def run(self):
        """主函数"""
        log.info('=' * 60)
        log.info('eBay 多线程商品详情爬虫')
        log.info('模式: 生产者保存HTML -> 队列传递文件路径 -> 消费者读取解析 -> 即用即删')
        log.info('=' * 60)

        # 初始化
        config = Config()

        storage = Storage_1(config)

        try:
            url_list = storage.read_input()
            print(f'读取到 {len(url_list)} 个URL')
        except FileNotFoundError as e:
            print(f'错误: {e}')
            return

        fetcher = Fetcher_1(config)
        crawler = MultiThreadCrawler_1(config, fetcher, storage)

        try:
            results = crawler.run(url_list)
            print(f'\n 爬取完成！共获取 {len(results)} 条数据')
        finally:
            fetcher.close()
            print('浏览器已关闭')

