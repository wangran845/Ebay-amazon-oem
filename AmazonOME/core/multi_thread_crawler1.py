"""
多线程爬虫协调器
"""
from __future__ import annotations
import threading
from typing import List, Dict, Tuple
from openpyxl import Workbook

from core.page_queue import PageQueue
from workers.producer import Producer,Producer_1
from workers.consumers import FullParserConsumer,FullParserConsumer_1
from utils.logger import get_logger
log = get_logger("MULTI")
class MultiThreadCrawler:
    """多线程爬虫主控"""

    def __init__(self, config1: 'Config', fetcher1: 'Fetcher', storage1: 'Storage'):
        self.config = config1
        self.fetcher = fetcher1
        self.storage = storage1

        self._results: List[Dict] = []
        self._result_lock = threading.Lock()

        self.stats = {
            'total': 0,
            'produced': 0,
            'consumed': 0,
            'failed': 0
        }

    def _save_result(self, result: Dict) -> None:
        """线程安全保存结果"""
        with self._result_lock:
            self._results.append(result)
            self.stats['consumed'] += 1

            if len(self._results) % 10 == 0:
                self._quick_save()

    def _quick_save(self) -> None:
        """快速保存临时结果"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(self.config.OUTPUT_HEADERS)

            for r in self._results:
                ws.append([
                    r.get('URL', ''),
                    r.get('OME_No', '')
                ])

            wb.save(self.config.TEMP_FILE)
            print(f'[Crawler]临时保存{len(self._results)}条')
        except Exception as e:
            print(f'[Crawler]临时保存失败: {e}')

    def run(self, url_list: List[Tuple[str, str, str]]) -> List[Dict]:
        """执行多线程爬取"""
        self.stats['total'] = len(url_list)
        print(f'[Crawler] 启动多线程爬虫，共 {self.stats["total"]} 个URL')
        print(f'[Crawler] 配置: {self.config.PRODUCER_NUM}生产者, {self.config.CONSUMER_NUM}消费者')

        # 创建队列
        page_queue = PageQueue(max_size=self.config.QUEUE_MAX_SIZE)

        # 创建生产者
        producer = Producer(
            config1=self.config,
            fetcher1=self.fetcher,
            page_queue=page_queue,
            url_list=url_list
        )

        # 创建消费者
        consumers = []
        for i in range(self.config.CONSUMER_NUM):
            consumer = FullParserConsumer(page_queue, self._save_result)
            consumers.append(consumer)

        # 启动所有线程
        print('[Crawler] 启动所有线程...')
        for c in consumers:
            c.start()
        producer.start()

        # 等待生产者完成
        producer.join()
        self.stats['produced'] = producer.success
        print(f'[Crawler] 生产者已结束，成功获取 {producer.success} 个页面')

        # 发送毒丸
        # 6. 等待队列清空后再发送毒丸
        print('[Crawler] 等待队列清空...')
        page_queue.join()  # 等待所有任务被处理完
        print('[Crawler] 发送毒丸信号给消费者...')
        page_queue.send_poison(len(consumers))

        # 等待消费者完成
        for c in consumers:
            c.join()
            print(f'[Crawler] {c.name} 已结束，处理了 {c.processed} 个任务')

        # 最终保存
        print(f'[Crawler] 全部完成，共收集 {len(self._results)} 条结果')
        self._final_save()

        return self._results

    def _final_save(self) -> None:
        """保存最终结果"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(self.config.OUTPUT_HEADERS)

            sorted_results = sorted(self._results, key=lambda x: x.get('_idx', 0))

            for r in sorted_results:
                ws.append([
                    r.get('URL', ''),
                    r.get('OME_No', '')
                ])

            self.config.ensure_dirs()
            wb.save(self.config.output_file)
            print(f'[Crawler]结果已保存: {self.config.output_file}')

        except Exception as e:
            print(f'[Crawler]最终保存失败: {e}')
            import traceback
            traceback.print_exc()

class MultiThreadCrawler_1:
    """多线程爬虫主控"""

    def __init__(self, config: 'Config', fetcher: 'Fetcher_1', storage: 'Storage_1'):
        self.config = config
        self.fetcher = fetcher
        self.storage = storage

        self._results: List[Dict] = []
        self._result_lock = threading.Lock()

        self.stats = {
            'total': 0,
            'produced': 0,
            'consumed': 0,
            'failed': 0
        }

    def _save_result(self, result: Dict) -> None:
        """线程安全保存结果"""
        with self._result_lock:
            self._results.append(result)
            self.stats['consumed'] += 1

            if len(self._results) % 10 == 0:
                self._quick_save()

    def _quick_save(self) -> None:
        """快速保存临时结果"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(self.config.OUTPUT_HEADERS_DE)

            for r in self._results:
                ws.append([
                    r.get('URL', ''),
                    r.get('OME_No', ''),
                    r.get('price', ''),
                    r.get("About_table", ''),
                    r.get('OEM', ''),
                    r.get('description_from_the_seller', ''),
                    r.get('seller_feedback', ''),
                    r.get('category', ''),
                    r.get('picture', ''),
                    r.get('OEM','')
                ])

            wb.save(self.config.TEMP_FILE_DE)
            print(f'[Crawler]临时保存{len(self._results)}条')
        except Exception as e:
            print(f'[Crawler]临时保存失败: {e}')

    def run(self, url_list: List[Tuple[str, str]]) -> List[Dict]:
        """执行多线程爬取"""
        self.stats['total'] = len(url_list)
        print(f'[Crawler] 启动多线程爬虫，共 {self.stats["total"]} 个URL')
        print(f'[Crawler] 配置: {self.config.PRODUCER_NUM}生产者, {self.config.CONSUMER_NUM}消费者')

        # 创建队列
        page_queue = PageQueue(max_size=self.config.QUEUE_MAX_SIZE)

        # 创建生产者
        producer = Producer_1(
            config=self.config,
            fetcher=self.fetcher,
            page_queue=page_queue,
            url_list=url_list
        )

        # 创建消费者
        consumers = []
        for i in range(self.config.CONSUMER_NUM):
            consumer = FullParserConsumer_1(page_queue, self._save_result)
            consumers.append(consumer)



        # 启动所有线程
        print('[Crawler] 启动所有线程...')
        for c in consumers:
            c.start()
        producer.start()

        # 等待生产者完成
        producer.join()
        self.stats['produced'] = producer.success
        print(f'[Crawler] 生产者已结束，成功获取 {producer.success} 个页面')

        # 发送毒丸
        # 6. 等待队列清空后再发送毒丸
        print('[Crawler] 等待队列清空...')
        page_queue.join()  # 等待所有任务被处理完
        print('[Crawler] 发送毒丸信号给消费者...')
        page_queue.send_poison(len(consumers))

        # 等待消费者完成
        for c in consumers:
            c.join()
            print(f'[Crawler] {c.name} 已结束，处理了 {c.processed} 个任务')

        # 最终保存
        print(f'[Crawler] 全部完成，共收集 {len(self._results)} 条结果')
        self._final_save()

        return self._results

    def _final_save(self) -> None:
        """保存最终结果"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.append(self.config.OUTPUT_HEADERS_DE)

            sorted_results = sorted(self._results, key=lambda x: x.get('_idx', 0))

            for r in sorted_results:
                ws.append([
                    r.get('URL', ''),
                    r.get('Title', ''),
                    r.get('price', ''),
                    r.get("About_table", ''),
                    r.get('item specifics', ''),
                    r.get('description_from_the_seller', ''),
                    r.get('seller_feedback', ''),
                    r.get('category', ''),
                    r.get('picture', ''),
                    r.get('OEM', '')
                ])

            self.config.ensure_dirs()
            wb.save(self.config.output_file_de)
            print(f'[Crawler]结果已保存: {self.config.output_file_de}')

        except Exception as e:
            print(f'[Crawler]最终保存失败: {e}')
            import traceback
            traceback.print_exc()