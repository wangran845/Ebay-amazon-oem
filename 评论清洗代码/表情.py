import pandas as pd
import re
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import logging
from datetime import datetime
import configparser
from pathlib import Path

# 尝试导入tqdm（可选依赖）
try:
    from tqdm import tqdm

    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False


# ==================== 配置管理 ====================

class Config:
    """配置管理类"""

    def __init__(self, config_file='excel_filter_config.ini'):
        self.config = configparser.ConfigParser(interpolation=None)
        self.config_file = config_file

        # 默认配置
        self.default_config = {
            'DEFAULT': {
                'allowed_chars': r'a-zA-Z0-9\s\.\,\!\?\;\:\'\"\(\)\[\]\{\}\<\>\/\@\#\$\%\^\&\*\-_\=\+\|\\\~\`',
                'output_suffix': '_filtered',
                'chunk_size': '10000',
                'show_progress': 'True',
                'use_fast_mode': 'True',
                'auto_backup': 'True'
            },
            'ADVANCED': {
                'max_file_size_mb': '100',
                'encoding': 'utf-8',
                'log_level': 'INFO'
            }
        }

        self.load_config()

    def load_config(self):
        """加载配置文件"""
        if os.path.exists(self.config_file):
            try:
                self.config.read(self.config_file, encoding='utf-8')
                print(f"已加载配置文件: {self.config_file}")
            except Exception as e:
                print(f"加载配置文件失败: {e}，使用默认配置")
                self.config.read_dict(self.default_config)
        else:
            self.config.read_dict(self.default_config)
            self.save_config()

    def save_config(self):
        """保存配置文件"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                self.config.write(f)
            print(f"配置文件已保存: {self.config_file}")
        except Exception as e:
            print(f"保存配置文件失败: {e}")

    def get_allowed_pattern(self):
        """获取允许的字符模式"""
        chars = self.config['DEFAULT']['allowed_chars']
        try:
            return re.compile(f'[^{chars}]')
        except re.error as e:
            print(f"正则表达式编译错误: {e}")
            # 使用简化版本作为后备
            return re.compile(r'[^a-zA-Z0-9\s\.\,\!\?\;\:\'\"\(\)\[\]\{\}\<\>\/\@\#\$\%\^\&\*\-_\=\+\|\\\~\`]')

    def get_chunk_size(self):
        """获取分块大小"""
        return int(self.config['DEFAULT']['chunk_size'])

    def use_fast_mode(self):
        """是否使用快速模式"""
        return self.config['DEFAULT'].getboolean('use_fast_mode')

    def auto_backup(self):
        """是否自动备份"""
        return self.config['DEFAULT'].getboolean('auto_backup')


# ==================== 日志管理 ====================

def setup_logging():
    """设置日志记录"""
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)

    log_file = log_dir / f"excel_filter_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)


# ==================== 字符过滤器 ====================

class CharacterFilter:
    """字符过滤器类（优化版）"""

    def __init__(self, config=None):
        self.config = config or Config()
        # 预编译正则表达式
        self.allowed_pattern = self.config.get_allowed_pattern()
        # 统计信息
        self.stats = {
            'total_cells': 0,
            'modified_cells': 0,
            'empty_cells': 0
        }

    def filter_text(self, text):
        """过滤文本"""
        if pd.isna(text):
            self.stats['empty_cells'] += 1
            return text

        original = str(text)
        try:
            filtered = self.allowed_pattern.sub('', original)
        except Exception as e:
            # 如果正则表达式出错，使用简单版本
            filtered = re.sub(r'[^\x00-\x7F]+', '', original)  # 只保留ASCII字符
            print(f"过滤出错，使用ASCII模式: {e}")

        if filtered != original:
            self.stats['modified_cells'] += 1

        return filtered

    def filter_dataframe(self, df, show_progress=False):
        """过滤整个DataFrame（使用map方法，兼容新版pandas）"""
        self.stats['total_cells'] = df.size

        # 创建过滤函数
        def filter_cell(x):
            return self.filter_text(x) if pd.notna(x) else x

        if show_progress and HAS_TQDM:
            # 使用tqdm显示进度
            result_df = df.copy()
            total_cells = df.size
            processed = 0

            for col in df.columns:
                for idx in df.index:
                    result_df.at[idx, col] = filter_cell(df.at[idx, col])
                    processed += 1
                    if processed % 1000 == 0:
                        tqdm.write(f"进度: {processed}/{total_cells} ({processed / total_cells * 100:.1f}%)")

            return result_df
        else:
            # 使用applymap（兼容处理）
            try:
                # 尝试使用map方法（pandas 2.1+）
                return df.map(filter_cell)
            except AttributeError:
                try:
                    # 尝试使用applymap（pandas 1.x）
                    return df.applymap(filter_cell)
                except AttributeError:
                    # 降级到逐列处理
                    result_df = df.copy()
                    for col in df.columns:
                        result_df[col] = df[col].apply(filter_cell)
                    return result_df

    def get_stats(self):
        """获取统计信息"""
        return self.stats

    def reset_stats(self):
        """重置统计信息"""
        self.stats = {
            'total_cells': 0,
            'modified_cells': 0,
            'empty_cells': 0
        }


# ==================== Excel处理函数 ====================

def create_backup(file_path):
    """创建文件备份"""
    try:
        backup_path = f"{file_path}.backup"
        if not os.path.exists(backup_path):
            import shutil
            shutil.copy2(file_path, backup_path)
            print(f"已创建备份: {backup_path}")
            return True
    except Exception as e:
        print(f"创建备份失败: {e}")
        return False


def process_excel_fast(file_path, sheet_name=0, output_file=None, config=None):
    """
    快速处理Excel文件（向量化操作）

    参数:
        file_path: Excel文件路径
        sheet_name: 工作表名称或索引
        output_file: 输出文件路径
        config: 配置对象
    """
    logger = logging.getLogger(__name__)
    config = config or Config()

    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return None

        # 自动备份
        if config.auto_backup():
            create_backup(file_path)

        # 读取Excel文件
        logger.info(f"正在读取文件: {file_path}")
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        logger.info(f"文件读取成功，共 {df.shape[0]} 行，{df.shape[1]} 列，{df.size} 个单元格")

        # 创建过滤器并处理
        filter_obj = CharacterFilter(config)

        # 使用向量化操作处理所有单元格
        logger.info("正在处理单元格...")
        df_filtered = filter_obj.filter_dataframe(df)

        # 生成输出文件名
        if output_file is None:
            base_name = os.path.splitext(file_path)[0]
            suffix = config.config['DEFAULT']['output_suffix']
            output_file = f"{base_name}{suffix}.xlsx"

        # 确保输出目录存在
        os.makedirs(os.path.dirname(os.path.abspath(output_file)) or '.', exist_ok=True)

        # 保存结果
        logger.info(f"正在保存结果到: {output_file}")
        df_filtered.to_excel(output_file, index=False, engine='openpyxl')

        # 输出统计信息
        stats = filter_obj.get_stats()
        logger.info(f"处理完成！")
        logger.info(f"总单元格数: {stats['total_cells']}")
        logger.info(f"修改的单元格数: {stats['modified_cells']}")
        logger.info(f"空单元格数: {stats['empty_cells']}")
        logger.info(f"输出文件: {output_file}")

        return df_filtered

    except Exception as e:
        logger.error(f"处理过程中出现错误: {e}", exc_info=True)
        return None


def process_excel_chunked(file_path, sheet_name=0, output_file=None, config=None):
    """
    分块处理大型Excel文件（适用于超大文件）

    参数:
        file_path: Excel文件路径
        sheet_name: 工作表名称或索引
        output_file: 输出文件路径
        config: 配置对象
    """
    logger = logging.getLogger(__name__)
    config = config or Config()
    chunk_size = config.get_chunk_size()

    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return None

        # 自动备份
        if config.auto_backup():
            create_backup(file_path)

        # 生成输出文件名
        if output_file is None:
            base_name = os.path.splitext(file_path)[0]
            suffix = config.config['DEFAULT']['output_suffix']
            output_file = f"{base_name}{suffix}.xlsx"

        # 确保输出目录存在
        os.makedirs(os.path.dirname(os.path.abspath(output_file)) or '.', exist_ok=True)

        logger.info(f"开始分块处理文件: {file_path}")
        logger.info(f"分块大小: {chunk_size} 行")

        filter_obj = CharacterFilter(config)
        total_rows = 0
        chunk_num = 0
        first_chunk = True

        # 分块读取和处理
        for chunk_num, chunk in enumerate(pd.read_excel(file_path, sheet_name=sheet_name, chunksize=chunk_size)):
            logger.info(f"处理第 {chunk_num + 1} 块，共 {len(chunk)} 行")

            # 处理当前块
            chunk_filtered = filter_obj.filter_dataframe(chunk)

            # 写入Excel
            if first_chunk:
                chunk_filtered.to_excel(output_file, index=False, engine='openpyxl')
                first_chunk = False
            else:
                # 追加模式
                try:
                    from openpyxl import load_workbook
                    book = load_workbook(output_file)
                    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                        writer.book = book
                        writer.sheets = {ws.title: ws for ws in book.worksheets}
                        chunk_filtered.to_excel(writer, index=False, header=False,
                                                startrow=writer.sheets['Sheet1'].max_row)
                except Exception as e:
                    logger.warning(f"追加写入失败，使用新文件: {e}")
                    chunk_filtered.to_excel(output_file, index=False, engine='openpyxl')

            total_rows += len(chunk)

        # 输出统计信息
        stats = filter_obj.get_stats()
        logger.info(f"分块处理完成！共处理 {chunk_num + 1} 块")
        logger.info(f"总行数: {total_rows}")
        logger.info(f"总单元格数: {stats['total_cells']}")
        logger.info(f"修改的单元格数: {stats['modified_cells']}")
        logger.info(f"输出文件: {output_file}")

        return True

    except Exception as e:
        logger.error(f"分块处理过程中出现错误: {e}", exc_info=True)
        return False


def process_excel_with_progress(file_path, sheet_name=0, output_file=None, config=None):
    """
    带进度条的处理函数
    """
    logger = logging.getLogger(__name__)
    config = config or Config()

    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            logger.error(f"文件不存在: {file_path}")
            return None

        # 自动备份
        if config.auto_backup():
            create_backup(file_path)

        # 读取Excel文件
        logger.info(f"正在读取文件: {file_path}")
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        logger.info(f"文件读取成功，共 {df.shape[0]} 行，{df.shape[1]} 列")

        filter_obj = CharacterFilter(config)

        # 使用进度条处理
        if HAS_TQDM:
            logger.info("使用tqdm进度条处理...")
            # 创建进度条
            total_cells = df.size
            with tqdm(total=total_cells, desc="处理单元格", unit="个") as pbar:
                def filter_with_progress(x):
                    result = filter_obj.filter_text(x) if pd.notna(x) else x
                    pbar.update(1)
                    return result

                try:
                    df_filtered = df.map(filter_with_progress)
                except AttributeError:
                    try:
                        df_filtered = df.applymap(filter_with_progress)
                    except AttributeError:
                        df_filtered = df.copy()
                        for col in df.columns:
                            df_filtered[col] = df[col].apply(filter_with_progress)
        else:
            logger.info("未安装tqdm，使用普通模式处理...")
            df_filtered = filter_obj.filter_dataframe(df)

        # 生成输出文件名
        if output_file is None:
            base_name = os.path.splitext(file_path)[0]
            suffix = config.config['DEFAULT']['output_suffix']
            output_file = f"{base_name}{suffix}.xlsx"

        # 保存结果
        logger.info(f"正在保存结果到: {output_file}")
        df_filtered.to_excel(output_file, index=False, engine='openpyxl')

        # 输出统计信息
        stats = filter_obj.get_stats()
        logger.info(f"处理完成！共处理 {stats['total_cells']} 个单元格")
        logger.info(f"输出文件: {output_file}")

        return df_filtered

    except Exception as e:
        logger.error(f"处理过程中出现错误: {e}", exc_info=True)
        return None


# ==================== 文件选择函数 ====================

def select_file_gui():
    """
    使用图形界面选择Excel文件
    """
    try:
        # 创建根窗口（隐藏）
        root = tk.Tk()
        root.withdraw()

        # 设置窗口置前
        root.attributes('-topmost', True)

        # 打开文件选择对话框
        file_path = filedialog.askopenfilename(
            title="请选择Excel文件",
            filetypes=[
                ("Excel文件", "*.xlsx *.xls"),
                ("所有文件", "*.*")
            ]
        )

        root.destroy()

        if file_path:
            return file_path
        else:
            print("未选择任何文件")
            return None

    except Exception as e:
        print(f"图形界面错误: {e}")
        return None


def select_file_cli():
    """
    命令行方式输入文件路径
    """
    while True:
        file_path = input("请输入Excel文件路径（输入 q 退出）: ").strip().strip('"').strip("'")

        if file_path.lower() == 'q':
            return None

        if os.path.exists(file_path):
            return file_path
        else:
            print(f"文件不存在: {file_path}")
            print("请重新输入正确的文件路径")


def validate_excel_file(file_path):
    """验证Excel文件是否有效"""
    try:
        # 检查文件扩展名
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            logging.warning(f"文件扩展名不是Excel格式: {file_path}")
            return False

        # 检查文件大小
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        config = Config()
        max_size = int(config.config['ADVANCED']['max_file_size_mb'])

        if file_size_mb > max_size:
            print(f"警告: 文件大小 {file_size_mb:.2f} MB 超过推荐大小 {max_size} MB")
            response = input("是否继续处理？(y/n): ").strip().lower()
            if response != 'y':
                return False

        # 尝试读取文件
        pd.ExcelFile(file_path)
        return True

    except Exception as e:
        logging.error(f"无效的Excel文件: {e}")
        return False


def select_sheet(file_path):
    """
    选择要处理的工作表
    """
    try:
        # 获取所有工作表名称
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names

        if len(sheet_names) == 1:
            print(f"文件只有一个工作表: {sheet_names[0]}")
            return sheet_names[0]

        print(f"\n文件包含以下工作表:")
        for i, name in enumerate(sheet_names, 1):
            print(f"  {i}. {name}")

        while True:
            try:
                choice = input(f"请选择要处理的工作表 (1-{len(sheet_names)}，直接回车使用第一个): ").strip()

                if choice == "":
                    return sheet_names[0]

                idx = int(choice) - 1
                if 0 <= idx < len(sheet_names):
                    return sheet_names[idx]
                else:
                    print(f"请输入 1-{len(sheet_names)} 之间的数字")
            except ValueError:
                print("请输入有效的数字")

    except Exception as e:
        print(f"读取工作表信息失败: {e}")
        return 0


# ==================== 报告生成 ====================

def generate_report(df, original_file, output_file, stats):
    """生成处理报告"""
    print("\n" + "=" * 60)
    print("处理报告")
    print("=" * 60)
    print(f"原始文件: {original_file}")
    if output_file:
        print(f"输出文件: {output_file}")
    print(f"处理后行数: {len(df)}")
    print(f"处理后列数: {len(df.columns)}")
    print(f"总单元格数: {stats['total_cells']}")
    print(f"修改的单元格数: {stats['modified_cells']}")
    if stats['total_cells'] > 0:
        print(f"修改比例: {stats['modified_cells'] / stats['total_cells'] * 100:.2f}%")
    print(f"空单元格数: {stats['empty_cells']}")
    print(f"列名: {', '.join(df.columns)}")

    # 保存报告到文件
    if output_file:
        report_file = f"{os.path.splitext(output_file)[0]}_report.txt"
    else:
        report_file = f"{os.path.splitext(original_file)[0]}_report.txt"

    try:
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("Excel内容筛选处理报告\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"原始文件: {original_file}\n")
            if output_file:
                f.write(f"输出文件: {output_file}\n")
            f.write(f"处理后行数: {len(df)}\n")
            f.write(f"处理后列数: {len(df.columns)}\n")
            f.write(f"总单元格数: {stats['total_cells']}\n")
            f.write(f"修改的单元格数: {stats['modified_cells']}\n")
            if stats['total_cells'] > 0:
                f.write(f"修改比例: {stats['modified_cells'] / stats['total_cells'] * 100:.2f}%\n")
            f.write(f"空单元格数: {stats['empty_cells']}\n")
            f.write(f"列名: {', '.join(df.columns)}\n")

        print(f"\n报告已保存到: {report_file}")
    except Exception as e:
        print(f"保存报告失败: {e}")


# ==================== 批量处理 ====================

def batch_process():
    """
    批量处理多个文件
    """
    logger = setup_logging()
    config = Config()

    print("\n" + "=" * 60)
    print("批量处理模式")
    print("=" * 60)

    # 选择文件列表
    file_paths = []

    print("请逐个输入Excel文件路径（输入空行结束）:")
    while True:
        file_path = input(f"文件 {len(file_paths) + 1}: ").strip().strip('"').strip("'")
        if not file_path:
            break
        if os.path.exists(file_path) and validate_excel_file(file_path):
            file_paths.append(file_path)
            print(f"  ✓ 已添加: {os.path.basename(file_path)}")
        else:
            print(f"  ✗ 文件无效，跳过: {file_path}")

    if not file_paths:
        print("未选择任何有效文件")
        return

    print(f"\n共选择 {len(file_paths)} 个文件")

    # 选择处理模式
    print("\n请选择处理模式:")
    print("1. 快速模式（推荐，适合普通文件）")
    print("2. 分块模式（适合超大文件）")
    print("3. 进度条模式（显示详细进度）")

    mode_choice = input("请选择 (1-3，默认1): ").strip()

    # 处理每个文件
    success_count = 0
    for i, file_path in enumerate(file_paths, 1):
        print(f"\n[{i}/{len(file_paths)}] 处理: {os.path.basename(file_path)}")
        print("-" * 40)

        # 选择工作表
        sheet_name = select_sheet(file_path)

        if mode_choice == '2':
            # 分块模式
            result = process_excel_chunked(file_path, sheet_name, config=config)
            if result:
                success_count += 1
        elif mode_choice == '3':
            # 进度条模式
            result = process_excel_with_progress(file_path, sheet_name, config=config)
            if result is not None:
                success_count += 1
        else:
            # 快速模式（默认）
            result = process_excel_fast(file_path, sheet_name, config=config)
            if result is not None:
                success_count += 1

    print(f"\n批量处理完成！成功: {success_count}/{len(file_paths)}")


# ==================== 配置管理界面 ====================

def manage_config():
    """配置管理"""
    config = Config()

    print("\n" + "=" * 60)
    print("配置管理")
    print("=" * 60)

    print("\n当前配置:")
    print(f"输出文件后缀: {config.config['DEFAULT']['output_suffix']}")
    print(f"分块大小: {config.config['DEFAULT']['chunk_size']} 行")
    print(f"快速模式: {config.config['DEFAULT']['use_fast_mode']}")
    print(f"自动备份: {config.config['DEFAULT']['auto_backup']}")
    print(f"最大文件大小: {config.config['ADVANCED']['max_file_size_mb']} MB")

    print("\n选项:")
    print("1. 重置为默认配置")
    print("2. 修改输出文件后缀")
    print("3. 修改分块大小")
    print("4. 切换快速模式")
    print("5. 切换自动备份")
    print("6. 返回主菜单")

    choice = input("\n请选择 (1-6): ").strip()

    if choice == '1':
        config.config.read_dict(config.default_config)
        config.save_config()
        print("配置已重置为默认值")

    elif choice == '2':
        new_suffix = input("请输入新的输出文件后缀（如 _filtered）: ").strip()
        if new_suffix:
            config.config['DEFAULT']['output_suffix'] = new_suffix
            config.save_config()
            print(f"输出文件后缀已改为: {new_suffix}")

    elif choice == '3':
        try:
            new_size = int(input("请输入新的分块大小（行数，如 5000）: ").strip())
            if new_size > 0:
                config.config['DEFAULT']['chunk_size'] = str(new_size)
                config.save_config()
                print(f"分块大小已改为: {new_size}")
        except ValueError:
            print("输入无效，请输入数字")

    elif choice == '4':
        current = config.config['DEFAULT'].getboolean('use_fast_mode')
        config.config['DEFAULT']['use_fast_mode'] = str(not current)
        config.save_config()
        print(f"快速模式已改为: {not current}")

    elif choice == '5':
        current = config.config['DEFAULT'].getboolean('auto_backup')
        config.config['DEFAULT']['auto_backup'] = str(not current)
        config.save_config()
        print(f"自动备份已改为: {not current}")


# ==================== 主函数 ====================

def main():
    """
    主函数 - 提供多种导入方式
    """
    # 设置日志
    logger = setup_logging()

    while True:
        print("\n" + "=" * 60)
        print("Excel 内容筛选工具 v2.0")
        print("功能：只保留数字、英文字母和常用标点符号")
        print("=" * 60)

        print("\n请选择功能:")
        print("1. 处理单个文件")
        print("2. 批量处理文件")
        print("3. 配置管理")
        print("4. 退出程序")

        choice = input("\n请输入选择 (1-4): ").strip()

        if choice == '1':
            # 处理单个文件
            print("\n请选择文件导入方式:")
            print("1. 图形界面选择文件")
            print("2. 手动输入文件路径")
            print("3. 返回上级菜单")

            sub_choice = input("请输入选择 (1-3): ").strip()

            if sub_choice == '1':
                file_path = select_file_gui()
            elif sub_choice == '2':
                file_path = select_file_cli()
            else:
                continue

            if not file_path:
                continue

            # 验证文件
            if not validate_excel_file(file_path):
                print("文件无效，请重新选择")
                continue

            print(f"\n已选择文件: {file_path}")

            # 选择工作表
            sheet_name = select_sheet(file_path)
            print(f"将处理工作表: {sheet_name}")

            # 询问输出文件路径
            print("\n输出文件设置:")
            use_default = input("是否使用默认输出文件名？(y/n, 默认y): ").strip().lower()

            output_file = None
            if use_default == 'n':
                output_file = input("请输入输出文件路径: ").strip().strip('"').strip("'")
                if not output_file:
                    output_file = None

            # 选择处理模式
            config = Config()
            print("\n请选择处理模式:")
            print("1. 快速模式（推荐，适合普通文件）")
            print("2. 分块模式（适合超大文件）")
            print("3. 进度条模式（显示详细进度）")

            mode_choice = input("请选择 (1-3，默认1): ").strip()

            # 开始处理
            print("\n开始处理...")

            if mode_choice == '2':
                # 分块模式
                result = process_excel_chunked(file_path, sheet_name, output_file, config)
                if result:
                    print("\n处理成功！")
            elif mode_choice == '3':
                # 进度条模式
                result = process_excel_with_progress(file_path, sheet_name, output_file, config)
                if result is not None:
                    print("\n处理成功！")
                    # 生成报告
                    generate_report(result, file_path, output_file, CharacterFilter(config).get_stats())
            else:
                # 快速模式（默认）
                result = process_excel_fast(file_path, sheet_name, output_file, config)
                if result is not None:
                    print("\n处理成功！")
                    # 生成报告
                    generate_report(result, file_path, output_file, CharacterFilter(config).get_stats())

            input("\n按回车键继续...")

        elif choice == '2':
            # 批量处理
            batch_process()
            input("\n按回车键继续...")

        elif choice == '3':
            # 配置管理
            manage_config()
            input("\n按回车键继续...")

        elif choice == '4':
            print("\n感谢使用，再见！")
            break

        else:
            print("无效的选择，请重新输入")


def quick_process():
    """
    快速处理 - 直接使用图形界面选择文件（快速模式）
    """
    file_path = select_file_gui()
    if file_path:
        if validate_excel_file(file_path):
            config = Config()
            result = process_excel_fast(file_path, config=config)
            if result is not None:
                print("\n处理成功！")
                generate_report(result, file_path, None, CharacterFilter(config).get_stats())
        else:
            print("文件无效")


if __name__ == "__main__":
    # 检查是否有命令行参数
    if len(sys.argv) > 1:
        # 命令行参数模式
        file_path = sys.argv[1]
        print(f"处理命令行指定的文件: {file_path}")
        if validate_excel_file(file_path):
            config = Config()
            result = process_excel_fast(file_path, config=config)
            if result is not None:
                generate_report(result, file_path, None, CharacterFilter(config).get_stats())
        else:
            print("文件无效")
    else:
        # 交互模式
        try:
            main()
        except KeyboardInterrupt:
            print("\n\n程序被用户中断")
        except Exception as e:
            logging.error(f"程序运行出错: {e}", exc_info=True)
            print(f"\n程序运行出错: {e}")
            input("按回车键退出...")